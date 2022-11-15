# -*- coding: utf-8 -*-
"""
Created on Tue Nov 15 16:12:57 2022

@author: xpess
"""


import os
import openpyxl
from pathlib import Path

## Paramètres 
dossier_comp = ""
fichier_comp = "CompetencesCPGE2021.xlsx"
onglet_comp  = "Competences"
filiere = 'PCSI-PSI' # Pour le choix du programme
discipline = "SII"
folder_path = "C:\\GitHub\ExercicesCompetences"


def read_file_competences(dossier_comp, fichier_comp,filiere,discipline) -> list:
    #### FROM EvaluationCompetences\scripts
    """
    Retourne la liste des competences contenu dans le fichier de competences
    Parameters
    ----------
    dossier_comp : TYPE
        DESCRIPTION.
    fichier_comp : TYPE
        DESCRIPTION.

    Returns
    -------
    list
        list(Competence).

    """
    # Lire un fichier de competences       
            
    xlsx_file = Path(dossier_comp, fichier_comp)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    
    # Read the active sheet:
    sheet = wb_obj.active
    nb_ligne = sheet.max_row
    #nb_col = sheet.max_column
    
    competences = []
    for row in sheet.iter_rows(max_row=nb_ligne):
        ligne = []
        for cell in row:
            ligne.append(cell.value)
        
        nb_none = 0
        for e in ligne : 
            if e==None :
                nb_none +=1
        if nb_none <= 2 :
            competences.append(ligne[0:2])
        
    return competences
    


def get_listes_exos(folder_path):
    """
    Réalise la liste des exos en retournant 
    rep, nom de fichier

    Parameters
    ----------
    folder_path : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    liste_exos = []
    for path, dirs, files in os.walk(folder_path):
        for filename in files : 
            if '.tex' in filename :
                liste_exos.append([path,filename])
    return liste_exos



competences = read_file_competences(dossier_comp, fichier_comp, filiere, discipline)
exercices = get_listes_exos(folder_path)

def entete(fid):
    
    fid.write("\\documentclass[10pt,fleqn]{book}\n")
    fid.write("\\usepackage[%\n")
    fid.write("\tpdftitle={Exercices de SII},\n")
    fid.write("\tpdfauthor={Xavier Pessoles}]{hyperref}\n")
    fid.write("\\newcommand{\\repRel}{../..}\n")
    fid.write("\\newcommand{\\repStyle}{\\repRel/Style}\n")
    fid.write("\\input{\\repRel/Style/packages}\n")
    fid.write("\\input{\\repRel/Style/new_style}\n")
    fid.write("\\input{\\repRel/Style/macros_SII}\n")
    fid.write("\\input{\\repRel/Style/environment}\n")
    fid.write("\\usepackage{\\repRel/Style/UPSTI_pedagogique}\n")

    fid.write("\\newcommand{\macrocomp}{macro_competences}\n")
    fid.write("\\newcommand{\comp}{competences}\n")
    fid.write("\\newcommand{\\td}{fichier_td}\n")
    fid.write("\\newcommand{\\repExos}{\\repRel/ExercicesCompetences}\n")
    fid.write("\\newcommand{\\repExo}{dossier}\n")

    fid.write("\\def\\xxYCartouche{-2.25cm}\n")
    fid.write("\\def\\xxYongletGarde{.5cm}\n")
    fid.write("\\def\\xxYOnget{.9cm}\n")

    fid.write("\\begin{document}\n")
    fid.write("\\def\\xxcompetences{}\n")
    fid.write("\\def\\xxfigures{}\n")
    fid.write("\\graphicspath{{\\repStyle/png/}}\n")

    fid.write("\\setlength{\columnseprule}{.1pt}\n")
    fid.write("\\input{\\repRel/Style/Entete_DDS}\n")

    fid.write("\\def\\xxpartie{}\n")
    fid.write("\\def\\xxnumpartie{}\n")
    fid.write("\\def\\xxchapitre{}\n")
    fid.write("\\def\\xxnumchapitre{}\n")
    fid.write("\\def\\xxactivite{DDS 3}\n")
    fid.write("\\def\\xxtitreexo{Les ptits devoirs du soir}\n")
    fid.write("\\def\\xxsourceexo{Xavier Pessoles}\n")

    fid.write("\\input{\\repRel/Style/pagegarde_TD}\n")

    fid.write("\\pagestyle{fancy}\n")
    fid.write("\\thispagestyle{plain}\n")

    fid.write("\\vspace{4.5cm}\n")

    fid.write("\\proffalse\n")



def write_tex(competences,exercices):
    codecomp = ""
    fid = open("test.tex","w",encoding="utf-8")
    entete(fid)

    
    for comp in competences : 
        
        if len(comp[0])==1 :## C'est une macro compétences >> Chapitre
            ligne = "\n\\chapter{"+comp[1]+"} \n"
            fid.write(ligne)
        elif len(comp[0])==2 :## C'est une compétence >> section
            ligne = "\n\\section{"+comp[1]+"} \n"
            fid.write(ligne)
            codecomp = comp[0]
        else  :## C'est une sous-comp >> subsection
            ligne = "\n\\subsection{"+comp[1]+"} \n"
            fid.write(ligne)
            code = comp[0]
            # On fait la liste des dossiers
            # On recherche si le code est présent
            
            os.chdir("C:/GitHub/ExercicesCompetences")
            rep = os.listdir()
            for r in rep :
                if codecomp in r :
                    code = code.replace("-","_")
                    
                    for exercice in exercices : 
                        if code in exercice[0]:
                            dossier = exercice[0][31:]
                            dossier = dossier.replace("\\","/")
                            exo = exercice[1]
                            
                            fid.write("\n\\renewcommand{\\repExo}{\\repExos/"+dossier+"}\n")
                            fid.write("\\renewcommand{\\td}{"+exo[:-4]+"}\n")
                            fid.write("\\graphicspath{{\\repStyle/png/}{\\repExo/images/}}\n")
                            fid.write("\\input{\\repExo/\\td.tex}\n")
                    #print()
                    a=1
    fid.write("\end{document} \n")        
    fid.close()
write_tex(competences,exercices)
            
        

